#! python3
# BendworksTextCombiner.py - A program to combine Greenlee bend reports into one line in an excel workbook for less printing and a cleaner format.

import os, sys, openpyxl, getpass, time, re, operator
from openpyxl.styles import colors, fonts, alignment, borders
from openpyxl.utils.cell import column_index_from_string
from openpyxl import worksheet
from datetime import datetime, date
from copy import copy
from fractions import Fraction as frac

def sortByIndex(List, index):
    List.sort(key = lambda i: i[index])
    return List

def stringToDec(strFrac):
    try:
        return float(strFrac)
    except ValueError:
        num, denom = strFrac.split('/')
        try:
            leading, num = num.split(' ')
            whole = float(leading)
        except ValueError:
            whole = 0
        frac = float(num) / float(denom)
        return whole - frac if whole < 0 else whole + frac

user = getpass.getuser()
currentYear = datetime.now().year
currentDate = date.today()

def decToString(dec):
    fraction = str(frac(dec))
    split = fraction.split('/')
    try:
        if int(split[0]) > int(split[1]):
            whole = int(split[0]) // int(split[1])
            numer = int(split[0]) - (whole * int(split[1]))
            return str(whole) + ' ' + str(numer) + '/' + str(split[1])
        else:
            return fraction
    except:
        return fraction

while True:
    try:
        os.chdir('V:\\1. VDC Projects\\GreenleeBendReports\\' + user + '\\ListsOfWantedBends')
        jobNum = str(input('Enter the job number:\n\n'))   # This is used to search the projects folder and get the specific file path name to save the combined file to.
        goodListName = input("\nEnter the name of the txt file with the Mark id's of wanted bend reports to combine. \n(This should be the exported schedule from Revit.)\n\n")
        startTime = time.time()
        goodListName = goodListName + '.txt'
        goodListFile = open(goodListName, 'rb').read()
        break
    except:
        print("\nFile couldn't be opened!\nCheck your spelling.\nMake sure your txt file is in the \"ListsOfWantedBends\" folder\n")

fileText = goodListFile.decode('utf16')     #Revit schedules need to be decoded to be able to read them.
wantedFiles = fileText.splitlines()

listOfFileNames = []
valuePairs = {}
for line in wantedFiles:
    cleanLine = line.replace('"', '')
    newLine = re.split('\t', cleanLine)
    listOfFileNames.append(newLine[0] + '.txt')
    try:
        valuePairs[newLine[0]] = newLine[1]
    except:
        pass

sortedValuePairs = sorted(valuePairs.items(), key=operator.itemgetter(1))

projectsDirPath = ('V:\\1. VDC Projects')
projectDirNames = os.listdir('V:\\1. VDC Projects')
targetDir = ('\\3- PFS\\GreenleeUsedBendReports')

jobFound = False
while not jobFound:
    if jobFound == True:
        break
    try:
        for directory in projectDirNames:
            if jobFound == True:
                break
            else:
                if jobNum in directory:
                    jobDirPath = (projectsDirPath + '\\' + directory)
                    jobName = str(jobDirPath.split(' - ')[1])
                    jobFound = True
        targetDir = jobDirPath + targetDir
    except:
        print('\nJob not found in VDC\Projects folder!\nRestartand double check to see if job number exists in the VDC projects folder.\n')
        exit(0)

# This portion creates the excel workbook, reads the wantedFiles values and appends the correct information to the workbook

bendWB = openpyxl.Workbook()
combinedSheet = bendWB.active
combinedSheet.title = 'Combined Bends'
labelSheet = bendWB.create_sheet('Labels')
reportSheet = bendWB.create_sheet('Report')

paramHeaderList = ['Conduit Type', 'Conduit Size', 'Conduit Group', 'Pipe Id', 'Num. of Bends', 'Cut Mark 1', 'Bend Marks', 'Cut Mark 2', 'Bend Angle', 'Bend Rotation', 'Conc. Bends', 'Error Code']
linesToRead = [6, 7, 10, 15, 29, 16, 30, 18, 17, 23, 33]
searchPath = ('V:\\1. VDC Projects\\GreenleeBendReports\\' + user + '\\BendWorksExports')
goodListName = goodListName.replace('.txt', '')
combinedFilePath = targetDir + '\\Combined' + goodListName.upper() + '.xlsx'

headerAlignment = alignment.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
headerFont = fonts.Font(size = 14, bold = True)

sheetHeaderList = {'Job Name': jobName, 'Job Number': jobNum, 'Area': goodListName, 'Submitted by': user}
colCounter = 1
for key, vals in sheetHeaderList.items():
    combinedSheet.cell(row = 1, column = colCounter).value = key
    combinedSheet.cell(row = 2, column = colCounter).value = vals
    combinedSheet.cell(row = 1, column = colCounter).alignment = headerAlignment
    combinedSheet.cell(row = 1, column = colCounter).font = headerFont
    combinedSheet.cell(row = 2, column = colCounter).alignment = headerAlignment
    combinedSheet.cell(row = 2, column = colCounter).font = headerFont
    colCounter += 1

# Bend Marks Merg Cells

combinedSheet.merge_cells('G3:K3')

colCounter = 1
for vals in paramHeaderList:
    if vals == 'Bend Marks':
        combinedSheet.cell(row = 3, column = 7).value = vals
        combinedSheet.cell(row = 3, column = 7).alignment = headerAlignment
        combinedSheet.cell(row = 3, column = 7).font = headerFont
        colCounter += 5
    else:
        combinedSheet.cell(row = 3, column = colCounter).value = vals
        combinedSheet.cell(row = 3, column = colCounter).alignment = headerAlignment
        combinedSheet.cell(row = 3, column = colCounter).font = headerFont
        colCounter +=1

filesFound = []
masterValues = []
rowCounter = 3
for file in os.listdir(searchPath):
    if file in listOfFileNames:
        filesFound.append(file)
        openFile = open(searchPath + '\\' + file).read()
        splitFile = openFile.splitlines()
        wantedValues = []        
        for num in linesToRead:
            value = splitFile[num]
            if num == 10:
                splitValue = value.split('       ')
                val2 = splitValue[1].strip()
                wantedValues.append(val2)            
            elif num == 16:
                splitValue = value.split(':')
                del splitValue[0]
                stringValue = ''
                for v in splitValue:
                    stringValue += v
                splitValue = stringValue.split('"')
                splitValue.pop()
                for v in splitValue:
                    val = v.strip()
                    wantedValues.append(val)            
            else:
                splitValue = value.split(':')
                val2 = splitValue[1].strip()
                wantedValues.append(val2)
        masterValues.append(wantedValues)

totalNumOfBends = str(len(masterValues))

namesToRemove = []
listsToRemove = []
concentricBends = []
for sublist in masterValues:  
    if sublist[3] == '0' and sublist[13] == '0':
        listsToRemove.append(masterValues.index(sublist))
        namesToRemove.append(sublist[2])
    elif sublist[3] == '0' and sublist[13] != '0':
        concentricBends.append(sublist[2])
for index in reversed(listsToRemove):
    masterValues.pop(index)

smallConduitList = []
bigConduitList = []
for strList in masterValues:
    newFloat = stringToDec(strList[1])
    strList[1] = newFloat
    if strList[2] in valuePairs.keys():
        strList.insert(2, valuePairs[strList[2]])
    if newFloat <= 2.0:
        smallConduitList.append(strList)
    elif newFloat >= 2.5:
        bigConduitList.append(strList)

smallSortedList = sortByIndex(smallConduitList, 1)
smallSortedList = sortByIndex(smallConduitList, 2)
bigSortedList = sortByIndex(bigConduitList, 2)
bigSortedList = sortByIndex(bigConduitList, 1)

sortedList = sortByIndex(masterValues, 1)

bendSize1Param = {.05: 0, .75: 0, 1.0: 0, 1.25: 0, 1.5: 0}  #For PFS tracking.
bendSize2Param = {2.0: 0, 2.5: 0}
bendSize3Param = {3.0: 0, 3.5: 0}
bendSize4Param = {4.0: 0}

emtConduit = {.05: 0, .75: 0, 1.0: 0, 1.25: 0, 1.5: 0, 2.0: 0, 2.5: 0, 3.0: 0, 3.5: 0, 4.0: 0}
rmcConduit = {.05: 0, .75: 0, 1.0: 0, 1.25: 0, 1.5: 0, 2.0: 0, 2.5: 0, 3.0: 0, 3.5: 0, 4.0: 0}
otherConduit = {.05: 0, .75: 0, 1.0: 0, 1.25: 0, 1.5: 0, 2.0: 0, 2.5: 0, 3.0: 0, 3.5: 0, 4.0: 0}

groupNames = []
for lists in sortedList:
    if lists[2] not in groupNames:
        groupNames.append(lists[2])

groupNames.sort()

for lists in sortedList:
    if lists[1] in bendSize1Param:
        bendSize1Param[lists[1]] += 1
        if lists[0] == 'EMT':
            emtConduit[lists[1]] += 1
        elif lists[0] == 'RIGID':
            rmcConduit[lists[1]] += 1
        else:
            otherConduit[lists[1]] += 1
    elif lists[1] in bendSize2Param:
        bendSize2Param[lists[1]] += 1
        if lists[0] == 'EMT':
            emtConduit[lists[1]] += 1
        elif lists[0] == 'RIGID':
            rmcConduit[lists[1]] += 1
        else:
            otherConduit[lists[1]] += 1
    elif lists[1] in bendSize3Param:
        bendSize3Param[lists[1]] += 1
        if lists[0] == 'EMT':
            emtConduit[lists[1]] += 1
        elif lists[0] == 'RIGID':
            rmcConduit[lists[1]] += 1
        else:
            otherConduit[lists[1]] += 1
    elif lists[1] in bendSize4Param:
        bendSize4Param[lists[1]] += 1
        if lists[0] == 'EMT':
            emtConduit[lists[1]] += 1
        elif lists[0] == 'RIGID':
            rmcConduit[lists[1]] += 1
        else:
            otherConduit[lists[1]] += 1

# Report Sheet Output

conduitTypes = [emtConduit, rmcConduit, otherConduit]       #Material req filling out
materialTypes = ['EMT', 'IMC', 'OTHER']
reportRow = 1
reportCol = 1
reportSheetHeader = ['QTY', 'SIZE', 'TYPE']
for elem in reportSheetHeader:
    reportSheet.cell(column=reportCol, row=1, value=elem)
    reportCol += 1
for types in conduitTypes:
    for k, v in types.items():
        if types[k]:
            reportRow += 1
            reportCol = 0            
            material  = ''
            
            if conduitTypes[0]:
                material = materialTypes[0]
            elif conduitType[1]:
                material = materialTypes[1]
            elif conduitType[2]:
                material = materialTypes[2]
            values = [v, k, material]
            for value in values:                
                reportCol +=1
                reportSheet.cell(column=reportCol, row=reportRow, value=value)

reportRow += 2
reportCol = 1
reportSheet.cell(column=reportCol, row=reportRow, value='Concentric Bends')
for quantity in concentricBends:
    reportRow += 1
    reportSheet.cell(column=reportCol, row=reportRow, value=quantity)

reportRow += 2
reportCol = 1
reportSheet.cell(column=reportCol, row=reportRow, value='Bend Reports Removed')
for quantity in namesToRemove:
    reportRow += 1
    reportSheet.cell(column=reportCol, row=reportRow, value=quantity) 

reportRow += 2
reportCol = 1
reportSheet.cell(column=reportCol, row=reportRow, value='Files Not Found')
for name in listOfFileNames:
    if name not in filesFound:
        reportRow += 1
        reportSheet.cell(column=reportCol, row=reportRow, value=name)

# Combined Sheet output

for data in smallSortedList:
    rowCounter +=2
    colCounter = 1
    for elem in data:
        if elem is data[1]:
            stringFraction = decToString(elem)
            combinedSheet.cell(row = rowCounter, column = colCounter).value = stringFraction
        else:
            combinedSheet.cell(row = rowCounter, column = colCounter).value = str(elem)
        colCounter +=1

for data in bigSortedList:
    rowCounter += 2
    colCounter = 1
    for elem in data:
        if elem is data[1]:
            stringFraction = decToString(elem)
            combinedSheet.cell(row = rowCounter, column = colCounter).value = stringFraction
        else:
            combinedSheet.cell(row = rowCounter, column = colCounter).value = str(elem)
        colCounter += 1

# Label Sheet Output

labelColCounter = 1
labelTitles = ['JobName', 'JobNumber', 'Area', 'Type', 'Size', 'Group ID', 'Bend Id', 'Sheet', 'GROUPS']
for item in labelTitles:
    labelSheet.cell(column=labelColCounter, row=1, value=item)
    labelColCounter +=1

commonLabels = [jobName, jobNum, goodListName]
labelRowCounter = 1
for lists in smallSortedList:
    labelRowCounter += 1
    for number in range(len(labelTitles)):
        labelsColCounter = 1
        for items in commonLabels:
            labelSheet.cell(column=labelsColCounter, row=labelRowCounter, value=items)
            labelsColCounter += 1
        for data in lists[0:4]:
            labelSheet.cell(column=labelsColCounter, row=labelRowCounter, value=data)
            labelsColCounter += 1

for lists in bigSortedList:
    labelRowCounter += 1
    for number in range(len(labelTitles)):
        labelsColCounter = 1
        for items in commonLabels:
            labelSheet.cell(column=labelsColCounter, row=labelRowCounter, value=items)
            labelsColCounter += 1
        for data in lists[0:4]:
            labelSheet.cell(column=labelsColCounter, row=labelRowCounter, value=data)
            labelsColCounter += 1

labelRowCounter = 2    
for group in groupNames:
    labelSheet.cell(column=9, row=labelRowCounter, value=group)
    labelRowCounter += 1

# This section applies formatting after all the values are processed

for x in range(4):
    combinedSheet.row_dimensions[x].height = 45

colWidths = {'colWidt10' : ['F', 'G', 'H', 'I', 'J', 'O'], 'colWidth12' : ['E', 'N', 'K', 'L'], 'colWidth15' : ['A', 'B', 'C', 'D'], 'colWidth35' : ['M', 'N']}

for key, value in colWidths.items():
    if '10' in key:
        for item in value:
            combinedSheet.column_dimensions[item].width = 10
    if '12' in key:
        for item in value:
            combinedSheet.column_dimensions[item].width = 12
    if '15' in key:
        for item in value:
            combinedSheet.column_dimensions[item].width = 15
    if '35' in key:
        for item in value:
            combinedSheet.column_dimensions[item].width = 35

valueRange = combinedSheet.max_row + 1
pipeIdColor = fonts.Font(color = 'FF9900')
cutMarkColor = fonts.Font(color = 'FF0000')
bendMarkColor = fonts.Font(color = '0000FF')
bendRotColor = fonts.Font(color = 'FF1493')
bendAngleColor = fonts.Font(color = '008000')

for cell in range(5 , valueRange):
    combinedSheet.cell(row = cell, column = 4).font = pipeIdColor
    combinedSheet.cell(row = cell, column = 6).font = cutMarkColor    
    combinedSheet.cell(row = cell, column = 7).font = bendMarkColor
    combinedSheet.cell(row = cell, column = 8).font = bendMarkColor
    combinedSheet.cell(row = cell, column = 9).font = bendMarkColor
    combinedSheet.cell(row = cell, column = 10).font = bendMarkColor
    combinedSheet.cell(row = cell, column = 11).font = bendMarkColor
    combinedSheet.cell(row = cell, column = 12).font = cutMarkColor
    combinedSheet.cell(row = cell, column = 13).font = bendRotColor
    combinedSheet.cell(row = cell, column = 14).font = bendAngleColor

# Need: to format printing 10 11x17 landscape
# Need: to format border
# Need: to print header at the top of each 11x17 sheet
# Need: to do some math to output the offset. (if it is an offset or kick 90.)
# Need: to get rid of the first cut mark and subtract it from the other marks
# Need: to figure out orientation of the kwik fit unicouple
# Need: to add Sheet info to label if it's ever used

bendWB.save(combinedFilePath)
print('\nThe workbook is saved in the following locaton : \n\n' + combinedFilePath)

PFSFilePath = projectsDirPath + '\\' + str(currentYear) + ' PRE-FAB TRACKING.xlsx'
testFilePath = 'V:\\5. VDC - Training\\7. RESEARCH AND DEVELOPMENT\\' + str(currentYear) + ' PRE-FAB TRACKING TEST.xlsx'

PFSTrackingwb = openpyxl.load_workbook(PFSFilePath)    # Replace file path with - PFSFilePath
trackingSheet = PFSTrackingwb.active
goodListName = goodListName.upper() + ' CONDUIT BENDS'
columnValues = {'A': jobNum, 'B': goodListName, 'C': currentDate, 'D': user + '.py', 'E': 1, 'F': 'SHOP', 'T': sum(bendSize1Param.values()), 'U': sum(bendSize2Param.values()), 'V': sum(bendSize3Param.values()), 'W': sum(bendSize4Param.values())}

def endStatement():
    ending = ('\nProcessed ' + totalNumOfBends + ' bends in a whopping ' + str(round(endTime - startTime, 2)) + " seconds!!!\n\nDon't forget to check the report Sheet!")
    return ending

rowToWriteIn = 5
for line in trackingSheet.iter_rows(min_row = 5, max_row = trackingSheet.max_row, min_col =2, max_col = 2, values_only = True):   
    if goodListName in str(line):
        print('\nBend area already found in PFS Tracking workbook. No values added.\n\nLabels created! Ready for a mail merge!')
        PFSTrackingwb.save(PFSFilePath)     # Replace file path with - PFSFilePath
        endTime = time.time()
        print(endStatement())
        time.sleep(2)
        exit(0)
    elif line[0] is None:   
        columnCounter = 1
        trackingSheet.insert_rows(rowToWriteIn)
        for column in range(trackingSheet.min_column, trackingSheet.max_column):
            formattedCell = trackingSheet.cell(row = rowToWriteIn - 1, column = columnCounter)
            newCell = trackingSheet.cell(row = rowToWriteIn, column = columnCounter)
            if formattedCell.has_style:
                newCell._style = copy(formattedCell._style)
                columnCounter += 1            
        for col, trackVal in columnValues.items(): 
            trackingSheet.cell(row = rowToWriteIn, column = column_index_from_string(col)).value = trackVal                    
        break
    else:
        rowToWriteIn +=1



PFSTrackingwb.save(PFSFilePath)    # Replace file path with - PFSFilePath
print('\nValues added to the PFS tracking sheet!\n\nLabels created! Ready for a mail merge!')
endTime = time.time()
print(endStatement())
time.sleep(2)

# Inserting a new row needs to carry the sum range with it.





